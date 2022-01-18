#!/usr/bin/env python3


import os
import csv
# 读取excel工具包
# 安装方式 pip install openpyxl
from openpyxl import load_workbook


def writeToFile(msg, fileName):
    """
    写文件
    """
    # 写 w 写 w+ 读写 会删除原有内容  a 追加
    with open(file=fileName, mode='a', encoding='utf-8') as f:
        f.write(msg)


def readAllFile(fileName):
    """
    读取整个文件
    """
    # 读
    with open(file=fileName, mode='r', encoding='utf-8') as f:
        content = f.read()
        print(content)


def printFile(fileName):
    """
    按行读文件
    """
    # 读
    with open(file=fileName, mode='r', encoding='utf-8') as f:
        lines = f.readlines()
        for line in lines:
            print(line)


def readCsv(fileName):
    """
    读取 csv文件
    """
    print("read  csv %s" % fileName)
    with open(fileName) as f:
        res = csv.reader(f)
        for row in res:
            # print(type(key))
            # key.split('\t')
            colu = row[0].split("\t")
            print(colu[0], colu[1])


def readExcel(filenName):
    """
    读取excel文件
    """
    print("read excel %s" % filenName)
    wb = load_workbook(filenName)
    # 通过索引获取sheet
    sheets = wb.worksheets
    for sheet in sheets:
        print(sheet)
    sheet = wb['Sheet1']
    # 读取sheet中的数
    # 行
    nrows = sheet.max_row
    # 列
    ncols = sheet.max_column
    print(nrows, ncols)

    # 打印行数 要从 1 到 max +1
    for i in range(1, nrows + 1):
        print("行数:", i)
    # 获取某个单元格的值 行列都从 1 开始计数
    value = sheet.cell(1, 1).value
    print(value)


def removeFile(fileName):
    os.remove(fileName)


def testReadCsvAndExcel():
    """
    执行命令要在当前文件所在目录下
    """
    file = '/conf/t-csv.csv'
    print(os.getcwd())
    # readCsv(os.getcwd() + file)
    excel = '/conf/t-excel.xlsx'
    readExcel(os.getcwd() + excel)


def testReadAndWrite():
    file = "test.log"
    writeToFile("test\n", fileName=file)
    writeToFile("test", fileName=file)
    writeToFile("test", fileName=file)
    # printFile(fileName=file)
    readAllFile(fileName=file)
    removeFile(file)


def run():
    print("start ...")
    testReadAndWrite()
    # testReadCsvAndExcel()
    print("end ...")


if __name__ == '__main__':
    """
    main 运行入口
    python3 py-string.py
    """
    run()
