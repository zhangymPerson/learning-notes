#!/usr/bin/env python3


import os
import csv

# 读取excel工具包
# 安装方式 pip install openpyxl
from openpyxl import load_workbook
import logging
# 配置日志
logging.basicConfig(level=logging.INFO,
                    format='[%(asctime)s-%(name)s-%(funcName)s-%(lineno)d-%(levelname)s]%(message)s',
                    datefmt='%Y-%m-%d %H:%M:%S')
logger = logging.getLogger(__name__)


def get_file_path():
    # 当前执行文件的路径，如 D：\aaa\bbb\ccc.py
    filepath = os.path.abspath(__file__)
    logger.info("%s,%s,%s", filepath, "a", "a")
    # 当前执行文件的上级路径，如 D：\aaa\bbb
    filepath = os.path.dirname(os.path.abspath(__file__))
    logger.info(filepath)
    # 继续向上
    filepath = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    logger.info(filepath)
    # 添加路径：
    # sys.path.append
    # logger.info(filepath)
    # 路径不存在则自动创建
    path = filepath + "/a/a.txt"
    logger.info(path)
    os.makedirs(path, exist_ok=True)
    # 路径组合
    # os.path.join(path1, path2, path3)
    logger.info(filepath)


def write_to_file(msg, file_name):
    """
    写文件
    """
    # 写 w 写 w+ 读写 会删除原有内容  a 追加
    with open(file=file_name, mode='a', encoding='utf-8') as f:
        f.write(msg)


def read_all_file(file_name):
    """
    读取整个文件
    """
    # 读
    with open(file=file_name, mode='r', encoding='utf-8') as f:
        content = f.read()
        logger.info(content)


def read_file(file_name):
    """
    按行读文件
    """
    # 读
    with open(file=file_name, mode='r', encoding='utf-8') as f:
        lines = f.readlines()
        for line in lines:
            # 去掉空字符
            line = line.strip()
            logger.info(line)


def get_file_word(file_name):
    """
    获取一个文件的内容 
    Args:
        params:fileName 文件名
    Returns:
        return 按行分割的内容集合
    Raises:
        列出与接口有关的所有异常.
    """
    with open(file_name, 'r', encoding='utf-8') as f:
        lines = f.readlines()
        list = []
        for line in lines:
            logger.info(line)
            list.append(line.rstrip())
        return list


def read_csv(file_name):
    """
    读取 csv文件
    """
    logger.info("file_name = [%s]" % (str(file_name)))
    with open(file_name) as f:
        res = csv.reader(f)
        for row in res:
            # logger.info(type(key))
            # key.split('\t')
            colu = row[0].split("\t")
            logger.info(colu[0], colu[1])


def read_excel(file_name):
    """
    读取excel文件
    """
    logger.info("read excel %s", file_name)
    wb = load_workbook(file_name)
    # 通过索引获取sheet
    sheets = wb.worksheets
    for sheet in sheets:
        logger.info(sheet)
    sheet = wb['Sheet1']
    # 读取sheet中的数
    # 行
    nrows = sheet.max_row
    # 列
    ncols = sheet.max_column
    logger.info(nrows, ncols)

    # 打印行数 要从 1 到 max +1
    for i in range(1, nrows + 1):
        logger.info("行数:%s" % i)
    # 获取某个单元格的值 行列都从 1 开始计数
    value = sheet.cell(1, 1).value
    logger.info(value)


def remove_file(file_name):
    """
    删除文件
    """
    os.remove(file_name)
    logger.info("删除文件 file_name = [%s]", file_name)



def test_read_csv_excel():
    """
    执行命令要在当前文件所在目录下
    """
    file = '/conf/t-csv.csv'
    logger.info(os.getcwd())
    # readCsv(os.getcwd() + file)
    excel = '/conf/t-excel.xlsx'
    read_excel(os.getcwd() + excel)


def test_read_write():
    """
    测试读写
    """
    file = "test.log"
    write_to_file("test\n", file_name=file)
    write_to_file("test", file_name=file)
    write_to_file("test", file_name=file)
    read_file(file_name=file)
    read_all_file(file_name=file)
    remove_file(file)


def run():
    """
    测试
    """
    logger.info("start ...")
    test_read_write()
    # testReadCsvAndExcel()
    get_file_path()
    logger.info("end ...")


if __name__ == '__main__':
    """
    main 运行入口
    python3 py-string.py
    """
    run()
